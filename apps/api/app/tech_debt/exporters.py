"""Deliverable renderers - turn a capability list into XLSX + PDF bytes.

Master Spec §15 Phase 3: "PDF + XLSX exporters for the deliverable."

XLSX: openpyxl. Header row + one row per capability + a summary row at
the bottom (Total Cost, Estimated Savings).

PDF: ReportLab. Pure Python; no native deps required (unlike WeasyPrint).
Phase 6 polish can revisit visual fidelity, but for v1 the deliverable is
a real, legitimate PDF with a title, summary, table, and savings figure.

Both renderers are pure functions over the data; no DB, no I/O. The
route layer writes the bytes via the existing StorageBackend.
"""

from __future__ import annotations

import io
from collections.abc import Iterable
from dataclasses import dataclass

from app.models.capability import CapabilityDisposition, CapabilityItem, CapabilityList


@dataclass(frozen=True)
class DeliverableContext:
    """Inputs the renderers share. Built once by the route layer."""

    client_legal_name: str
    service_title: str
    cap_list: CapabilityList
    items: list[CapabilityItem]
    total_cost: float
    estimated_savings: float
    savings_cost_known: bool
    # Reconciliation (N-010). The workspace disclosed "N rows received · M
    # included · K excluded"; the deliverable did not, so a released report
    # stated an unqualified "Total annual cost" for an upload whose excluded
    # rows were worth $240,000. Defaulted so older callers still construct.
    source_rows_total: int | None = None
    excluded_count: int = 0
    # Source-derived items only — children of a decomposed bundle are excluded,
    # so splitting can never move the reconciliation arithmetic.
    included_count: int = 0
    # False when the excluded rows exist but could not be named individually.
    # The count is exact either way; only the naming is withheld.
    excluded_rows_named: bool = False
    # False when an INCLUDED item carries no cost, so `total_cost` is a floor.
    #
    # The twin of `savings_cost_known`, which has existed since this dataclass
    # did. #126 asked the floor question of SPEND and was answered on the
    # dashboard (`spend_completeness`) and not here, so for one commit the
    # dashboard called a list "partial" while the released document called the
    # same list's figure a "Total". Two surfaces, one list, contradictory
    # claims, and the client only keeps the document.
    #
    # Defaulted True so older callers construct unchanged: a context built
    # without this field is one whose items were never checked, and the
    # pre-existing behaviour for those is the unqualified label.
    spend_cost_known: bool = True


def reconciliation_line(ctx: DeliverableContext) -> str | None:
    """ "N rows received · M included · K excluded", or None when nothing was dropped.

    A client uploaded 28 rows worth $3,608,000; two were excluded and the
    released PDF/XLSX/DOCX said only "Total annual cost: $3,368,000". The report
    has to carry the same reconciliation the workspace shows, or it states a
    partial figure as a total.
    """
    if not ctx.excluded_count or ctx.source_rows_total is None:
        return None
    line = (
        f"{ctx.source_rows_total} rows received · {ctx.included_count} included · "
        f"{ctx.excluded_count} excluded"
    )
    if not ctx.excluded_rows_named:
        # A count with no accompanying list reads as a rendering bug unless it
        # says why. `reconcile.py` withholds the names rather than guessing when
        # the provider did not attribute every item to a source row; the count is
        # still exact, and saying so is what keeps it usable.
        line += " (excluded rows were not attributed individually)"
    return line


def cost_label(ctx: DeliverableContext) -> str:
    """Never call a partial figure a total (UX finding #4).

    "Total annual cost" is the ONLY label that asserts completeness, and it is
    reached only when nothing was excluded, the reconciliation is on record,
    AND every included item carries a cost. Each of those was a separate way
    for the old two-valued label to overstate:

      - "nothing was excluded" vs "whether anything was excluded was never
        recorded" -- different claims about the upload, and for the whole life
        of this function they shared the string "Total annual cost".
      - "every item is costed" vs "some item is not" -- `total_cost` skips an
        uncosted item, so the figure is a floor and nothing said so. That half
        was fixed on the dashboard first and here second, which for one commit
        had the two surfaces contradicting each other about the same list.

    The hole was `source_rows_total is None`. `build_context` derives
    `excluded_count = max(received - included, 0) if received is not None else 0`,
    so a list with no reconciliation on record yields 0, and 0 read as "nothing
    was excluded". `reconciliation_line` -- the function immediately above --
    guards that exact condition at its first line and correctly returns None.
    Two adjacent functions, one predicate each, and only one of them had the
    guard: the report printed no reconciliation line, which is right, and then
    called the figure a TOTAL, which is not.

    Who has a NULL `source_rows_total`: pre-0036 lists, and any list not cut by
    an AI extraction (`reconcile.py` types `received` as a plain `int`, so
    `routes/tech_debt.py`'s `source_rows_total=result.reconciliation.received`
    always writes one when extraction ran). `services/stages.py`, in the helper
    returning `getattr(cap_list, "source_rows_total", None) is not None`,
    already reads NULL here as "un-analysed rather than
    analysed", calling that "the conservative direction" -- this makes the
    renderer agree with that reading instead of contradicting it.

    NOT the same defect as the one `build_context`'s comment already records.
    That one is `received is not None` with items >= source rows, where the
    subtraction floors to 0 while rows genuinely were excluded, and it needs
    `attribution_complete` persisted to fix. This one needs two lines, and the
    existing comment is silent on it -- which is why it survived: the comment
    sits exactly where a reader would check and describes a NARROWER case than
    they would assume it covers.
    """
    if ctx.source_rows_total is not None and ctx.included_count > ctx.source_rows_total:
        # THE UNBALANCED CASE. More items than there were source rows, so
        # `excluded_count` floored to 0 and the two branches below both fall
        # through to "Total annual cost" -- a completeness claim over an
        # accounting that cannot balance.
        #
        # This was deliberately left for one commit, on the stated reason that
        # naming the fault needs `reconcile.py`'s `attribution_complete`
        # persisted. That reason was FALSE and the same commit disproved it:
        # `tech_debt_dashboard` withholds the identical claim using these two
        # operands and no migration, and `DeliverableContext` has carried both
        # since `build_context` set them. The migration buys the ability to say
        # WHY -- "the reconciliation does not balance" rather than "may not be
        # complete" -- not the ability to stop overstating.
        #
        # Withholding the claim needs no migration, so it is not deferred. The
        # dashboard and the document now agree on this list; for one commit
        # they did not, which is the defect #126 exists to end, reinstated in a
        # state nobody had listed.
        return "Annual cost (may not be complete)"
    if not ctx.spend_cost_known:
        # Above `excluded_count`, and the precedence is deliberate. An uncosted
        # included item
        # makes the NUMBER a floor, which is a stronger qualification than
        # "some rows were excluded" -- "Included annual cost" would still
        # assert that what IS included was fully counted, and it was not.
        # Nothing is lost by not naming the exclusion here: `reconciliation_line`
        # states the received/included/excluded counts on its own.
        return "Annual cost (may not be complete)"
    if ctx.excluded_count:
        return "Included annual cost"
    if ctx.source_rows_total is None:
        return "Annual cost (may not be complete)"
    return "Total annual cost"


def _disposition_label(d: CapabilityDisposition | None) -> str:
    if d is None:
        return "Undecided"
    return {
        CapabilityDisposition.KEEP: "Keep",
        CapabilityDisposition.CONSOLIDATE: "Consolidate",
        CapabilityDisposition.CUT: "Cut",
    }[d]


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    cap_list: CapabilityList,
    items: Iterable[CapabilityItem],
) -> DeliverableContext:
    items_list = list(items)
    total_cost = 0.0
    estimated_savings = 0.0
    savings_known = True
    # `total_cost` SKIPS an uncosted item rather than failing, which makes the
    # figure a floor. Nothing recorded that, so `cost_label` had no way to know
    # and printed "Total annual cost" over it (#126, exporter half).
    spend_known = True
    for it in items_list:
        if it.annual_cost_usd is not None:
            total_cost += float(it.annual_cost_usd)
        else:
            spend_known = False
        if it.disposition == CapabilityDisposition.CUT:
            if it.annual_cost_usd is None:
                savings_known = False
            else:
                estimated_savings += float(it.annual_cost_usd)
    named = list(getattr(cap_list, "excluded_rows", None) or [])
    received = getattr(cap_list, "source_rows_total", None)
    # Rows that came from the upload. Children of a decomposed bundle carry a
    # parent and are NOT source rows — counting them made `28 > 32` false in the
    # workspace and unmounted the whole disclosure until 2026-08-07.
    included = sum(1 for it in items_list if getattr(it, "parent_item_id", None) is None)
    # Derive the count; do not measure the NAMED list. `reconcile.py` withholds
    # the names when the provider did not attribute every item to a source row,
    # so `len(named)` is 0 in exactly the case where rows WERE excluded and
    # nobody can say which — the 2026-08-04 defect, reachable through the
    # mechanism added to prevent it. It equals `len(named)` whenever attribution
    # is complete, so deriving it needs no second stored counter to drift on the
    # include-an-excluded-row path.
    #
    # NOT trustworthy in every regime, and an earlier version of this comment
    # claimed it was. When the model emits at least as many items as there were
    # source rows — two items sharing one `source_row_index`, say — `max(..., 0)`
    # reports ZERO, and a genuinely excluded row goes undisclosed. Strictly
    # better than measuring the named list, which was 0 in that case too; still
    # short of honest.
    #
    # NEITHER RENDERER OVERSTATES ANY MORE: the dashboard reports "partial" and
    # `cost_label` returns "Annual cost (may not be complete)" for this case,
    # both from these two counts and without a migration. What remains is
    # naming the CAUSE — "the reconciliation does not balance" rather than a
    # generic qualifier — which needs `reconcile.py`'s `attribution_complete`
    # persisted: a zero-value record that names the fault, per the CLAUDE.md
    # rule. Tracked as #193.
    #
    # The number is here because a disposition asserting it is tracked, with
    # nothing tracking it, is an unfixed defect wearing a managed one's
    # costume. Written as one account rather than a claim plus a correction
    # appended below it, because a reader going top-down would otherwise meet
    # the superseded sentence first.
    excluded_count = max(received - included, 0) if received is not None else 0
    return DeliverableContext(
        source_rows_total=received,
        excluded_count=excluded_count,
        included_count=included,
        excluded_rows_named=bool(named),
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        cap_list=cap_list,
        items=items_list,
        total_cost=total_cost,
        estimated_savings=estimated_savings,
        savings_cost_known=savings_known,
        spend_cost_known=spend_known,
    )


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: DeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl returned no active worksheet")
    ws.title = "Capability List"

    header = [
        "Name",
        "Vendor",
        "Category",
        "Function",
        "Annual Cost (USD)",
        "Licenses",
        "Disposition",
        "Rationale",
        "Notes",
        "AI Confidence %",
    ]
    ws.append(header)
    header_fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for item in ctx.items:
        ws.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                item.function or "",
                float(item.annual_cost_usd) if item.annual_cost_usd is not None else "",
                item.license_count if item.license_count is not None else "",
                _disposition_label(item.disposition),
                item.disposition_rationale or "",
                item.notes or "",
                item.confidence_pct if item.confidence_pct is not None else "",
            ]
        )

    # Summary row at the bottom.
    summary_row = ws.max_row + 2
    recon = reconciliation_line(ctx)
    if recon:
        ws.cell(row=summary_row, column=1, value="Reconciliation").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=recon)
        ws.cell(
            row=summary_row + 1,
            column=2,
            value="Figures below cover the included rows only, not the whole upload.",
        ).font = Font(italic=True)
        summary_row += 3
    ws.cell(row=summary_row, column=1, value=cost_label(ctx)).font = Font(bold=True)
    ws.cell(row=summary_row, column=5, value=ctx.total_cost).number_format = "$#,##0"
    ws.cell(row=summary_row + 1, column=1, value="Estimated annual savings").font = Font(bold=True)
    savings_cell = ws.cell(row=summary_row + 1, column=5, value=ctx.estimated_savings)
    savings_cell.number_format = "$#,##0"
    if not ctx.savings_cost_known:
        ws.cell(
            row=summary_row + 1,
            column=6,
            value="≥ (one or more cut rows missing a cost)",
        ).font = Font(italic=True)

    # Reasonable column widths.
    widths = [28, 22, 16, 28, 18, 10, 14, 38, 38, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_pdf(ctx: DeliverableContext) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=f"{ctx.service_title} — {ctx.client_legal_name}",
        author="SHIELD by Kentro",
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(Paragraph(ctx.service_title, h1))
    story.append(Paragraph(ctx.client_legal_name, body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Summary", h2))
    recon = reconciliation_line(ctx)
    if recon:
        story.append(Paragraph(f"<b>{recon}</b>", body))
        story.append(
            Paragraph(
                "Figures below cover the included rows only, not the whole upload.",
                body,
            )
        )
        story.append(Spacer(1, 0.08 * inch))
    savings = (
        f"${ctx.estimated_savings:,.0f}"
        if ctx.savings_cost_known
        else f"≥ ${ctx.estimated_savings:,.0f}"
    )
    story.append(
        Paragraph(
            f"Capabilities reviewed: <b>{len(ctx.items)}</b> · "
            f"{cost_label(ctx)}: <b>${ctx.total_cost:,.0f}</b> · "
            f"Estimated annual savings: <b>{savings}</b>",
            body,
        )
    )
    if not ctx.savings_cost_known:
        story.append(
            Paragraph(
                "Note: at least one row marked <i>Cut</i> is missing an annual cost. "
                "The savings figure is a lower bound.",
                body,
            )
        )

    story.append(Paragraph("Capability list", h2))

    table_data: list[list] = [["Name", "Vendor", "Category", "Annual cost", "Disposition"]]
    for item in ctx.items:
        cost = f"${float(item.annual_cost_usd):,.0f}" if item.annual_cost_usd is not None else "—"
        table_data.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                cost,
                _disposition_label(item.disposition),
            ]
        )

    table = Table(
        table_data,
        colWidths=[2.2 * inch, 1.4 * inch, 1.2 * inch, 1.0 * inch, 1.2 * inch],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0e1220")),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ]
        )
    )
    story.append(table)

    doc.build(story)
    return out.getvalue()


# ---------------------------------------------------------------------------
# DOCX (Work Order C4) - mirrors the PDF content.
# ---------------------------------------------------------------------------


def render_docx(ctx: DeliverableContext) -> bytes:
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(f"{ctx.service_title} — {ctx.client_legal_name}")
    add_title(doc, ctx.service_title, ctx.client_legal_name)

    savings = (
        f"${ctx.estimated_savings:,.0f}"
        if ctx.savings_cost_known
        else f"≥ ${ctx.estimated_savings:,.0f}"
    )
    add_heading(doc, "Summary")
    recon = reconciliation_line(ctx)
    lines = (
        [recon, "Figures below cover the included rows only, not the whole upload."]
        if recon
        else []
    ) + [
        f"Capabilities reviewed: {len(ctx.items)}",
        f"{cost_label(ctx)}: ${ctx.total_cost:,.0f}",
        f"Estimated annual savings: {savings}",
    ]
    if not ctx.savings_cost_known:
        lines.append(
            "Note: at least one row marked Cut is missing an annual cost. "
            "The savings figure is a lower bound."
        )
    add_paragraphs(doc, lines)

    add_heading(doc, "Capability list")
    rows = []
    for item in ctx.items:
        cost = f"${float(item.annual_cost_usd):,.0f}" if item.annual_cost_usd is not None else "—"
        rows.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                cost,
                _disposition_label(item.disposition),
            ]
        )
    add_table(doc, ["Name", "Vendor", "Category", "Annual cost", "Disposition"], rows)

    return to_bytes(doc)
