"""Exported documents must use the target the CONSULTANT approved against.

Issues #73 (ZT exporter drops both targets), #75 (ZT Gap Plan truncated with the
true total rendered nowhere) and #79 (the CSF twin — exporter and the /home
value card use the engine default while the client dashboard uses the client's
tier).

One test file because it is one defect in three places: `analyze_gaps` called
without the target the rest of the product resolves. The intake tier/stage is
MANDATORY (`routes/intake.py` 422s a request without one), so every real
engagement has one and roughly two-thirds of them are not the default.

Why no existing test caught it: `seed_demo.py` creates services with no
`source_request_id`, and every e2e spec creates its service by direct POST. Both
therefore get the default, and all surfaces agree. The fixtures could not
express the failure — another instance of #72.
"""

from __future__ import annotations

import os
import uuid
from collections.abc import Iterator
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.storage.local import LocalFilesystemStorage


@pytest.fixture()
def app_client(tmp_path) -> Iterator[TestClient]:
    url = f"sqlite:///{tmp_path / 'shield-export-targets.db'}"
    os.environ["DATABASE_URL"] = url
    api_root = Path(__file__).resolve().parents[2]
    cfg = Config(str(api_root / "alembic.ini"))
    cfg.set_main_option("script_location", str(api_root / "alembic"))
    cfg.set_main_option("sqlalchemy.url", url)
    command.upgrade(cfg, "head")
    engine = create_engine(url, future=True)
    TestSession = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    storage = LocalFilesystemStorage(tmp_path / "storage")

    from app.db.session import get_db
    from app.main import create_app
    from app.routes.artifacts import _storage_dep

    def override_get_db() -> Iterator[Session]:
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[_storage_dep] = lambda: storage

    from app.models.client import Client as _Client
    from app.models.client_domain import ClientDomain as _ClientDomain

    seed = TestSession()
    tenant = _Client(legal_name="Test Tenant")
    seed.add(tenant)
    seed.flush()
    seed.add(_ClientDomain(client_id=tenant.id, domain="example.com"))
    seed.commit()
    cid = str(tenant.id)
    seed.close()

    with TestClient(app, headers={"X-Client-Id": cid}) as c:
        yield c


def _register(c: TestClient, email: str) -> dict:
    r = c.post(
        "/auth/register",
        json={
            "email": email,
            "password": "correct horse battery staple!",
            "display_name": email.split("@")[0],
        },
    )
    assert r.status_code == 201, r.text
    return r.json()


def _attach_intake_target(svc_id: str, *, csf_tier: int | None = None, zt_stage: int | None = None):
    """Give the service a source request carrying the client's chosen target.

    This is what the seed and every e2e spec skip, and why the defect survived.
    """
    from app.models.service import Service as _Service
    from app.models.service_request import ServiceRequest as _SR

    eng = create_engine(os.environ["DATABASE_URL"], future=True)
    with sessionmaker(bind=eng, future=True)() as s:
        svc = s.get(_Service, uuid.UUID(svc_id))
        sr = _SR(
            client_id=svc.client_id,
            service_type=svc.kind.value,
            requested_by=svc.opened_by,
            csf_target_tier=csf_tier,
            zt_target_stage=zt_stage,
        )
        s.add(sr)
        s.flush()
        svc.source_request_id = sr.id
        s.commit()


# --- #73 / #75 — Zero Trust ------------------------------------------------


@pytest.mark.unit
def test_zt_export_uses_the_engagement_target_not_the_engine_default(app_client) -> None:
    """The gap set in the document must match the one the consultant approved.

    `finalize_zt_deliverable` called `analyze_gaps` with neither `targets` nor
    `target_stage`, so it fell back to DEFAULT_TARGET_STAGE (3) for every
    capability while `/gap-analysis` — the view the consultant approves from —
    used the client's stage. A client on stage 4 received a document listing a
    different set of gaps than the one that was signed off.
    """
    c = app_client
    admin = _register(c, "admin@example.com")
    h = {"Authorization": f"Bearer {admin['tokens']['access_token']}"}

    svc_id = c.post(
        "/zt/services", headers=h, json={"kind": "zero_trust_cisa", "title": "ZT"}
    ).json()["id"]
    _attach_intake_target(svc_id, zt_stage=4)

    a = c.post(f"/zt/services/{svc_id}/assessments", headers=h).json()
    # Stage 3 everywhere: ZERO gaps against the default target of 3, but every
    # capability is a gap against the client's chosen 4.
    for ans in a["answers"]:
        c.patch(f"/zt/answers/{ans['id']}", headers=h, json={"maturity_stage": 3})
    c.post(f"/zt/assessments/{a['id']}/approve", headers=h)

    approved = c.get(
        f"/zt/services/{svc_id}/gap-analysis", headers=h, params={"target_stage": 4}
    ).json()
    assert approved["total_gap_count"] > 0, "fixture must produce gaps at the client's target"

    fin = c.post(f"/zt/services/{svc_id}/deliverables/finalize", headers=h)
    assert fin.status_code == 201, fin.text

    from app.models.deliverable import Deliverable

    eng = create_engine(os.environ["DATABASE_URL"], future=True)
    with sessionmaker(bind=eng, future=True)() as s:
        deliv = s.get(Deliverable, uuid.UUID(fin.json()["id"]))
        summary = deliv.summary or ""

    # The summary line is the cheapest observable proof of which target the
    # document was built against. Assert the TARGET, not a bare gap count:
    # "37/37 capabilities scored" already contains "37", so
    # `str(total_gap_count) in summary` is satisfied by the coverage fraction
    # whether or not the fix is present. Both halves below name the target.
    assert (
        "at target S4" in summary
    ), f"deliverable summary was built against the default, not the client's target: {summary!r}"
    assert (
        f"{approved['total_gap_count']} gap(s) at target S4" in summary
    ), f"document reports a different gap count than the approved view: {summary!r}"


@pytest.mark.unit
def test_zt_export_states_the_true_gap_total_when_it_truncates(app_client) -> None:
    """#75: the Gap Plan SHEET renders a `top_n` slice and never states the total.

    Asserted on the rendered XLSX, not on the deliverable summary line. The
    summary already carries the true count, so asserting it there passes whether
    or not the defect exists — which is exactly the #72 shape, and the first
    draft of this test did it.

    The on-screen list discloses the total (`ZtGapList` renders
    `total_gap_count`); the client's document must not be the one surface that
    omits it.
    """
    import io

    from openpyxl import load_workbook

    c = app_client
    admin = _register(c, "admin@example.com")
    h = {"Authorization": f"Bearer {admin['tokens']['access_token']}"}

    svc_id = c.post(
        "/zt/services", headers=h, json={"kind": "zero_trust_cisa", "title": "ZT"}
    ).json()["id"]
    a = c.post(f"/zt/services/{svc_id}/assessments", headers=h).json()
    for ans in a["answers"]:
        c.patch(f"/zt/answers/{ans['id']}", headers=h, json={"maturity_stage": 1})
    c.post(f"/zt/assessments/{a['id']}/approve", headers=h)

    fin = c.post(f"/zt/services/{svc_id}/deliverables/finalize", headers=h)
    assert fin.status_code == 201, fin.text

    gaps = c.get(f"/zt/services/{svc_id}/gap-analysis", headers=h).json()
    total = gaps["total_gap_count"]
    assert total > 20, "fixture must actually exercise truncation, or this proves nothing"

    art_id = fin.json()["xlsx_artifact_id"]
    raw = c.get(f"/artifacts/{art_id}/download", headers=h).content
    ws = load_workbook(io.BytesIO(raw))["Gap Plan"]
    rendered = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    blob = " ".join(rendered)

    shown = ws.max_row - 2  # minus the caption row AND the header
    assert shown < total, f"fixture did not truncate: shown={shown} total={total}"
    assert str(total) in blob, (
        f"the Gap Plan shows {shown} of {total} gaps and never states the total — "
        "a client cannot tell anything was omitted"
    )


@pytest.mark.unit
def test_zt_export_honours_a_per_capability_target_override(app_client) -> None:
    """The OTHER half of #73: `targets=` was as absent as `target_stage=`.

    `test_zt_export_uses_the_engagement_target_not_the_engine_default` leaves
    every per-capability `target_stage` NULL, so `targets=targets_map` could be
    deleted from `finalize_zt_deliverable` and the whole suite stayed green —
    the exact case that comment cites ("a capability stored with target 2
    printed as 3") was the one nothing exercised.
    """
    c = app_client
    admin = _register(c, "admin@example.com")
    h = {"Authorization": f"Bearer {admin['tokens']['access_token']}"}

    svc_id = c.post(
        "/zt/services", headers=h, json={"kind": "zero_trust_cisa", "title": "ZT"}
    ).json()["id"]
    _attach_intake_target(svc_id, zt_stage=4)

    a = c.post(f"/zt/services/{svc_id}/assessments", headers=h).json()
    answers = a["answers"]
    for ans in answers:
        c.patch(f"/zt/answers/{ans['id']}", headers=h, json={"maturity_stage": 3})
    # One capability the consultant has agreed to leave at stage 3. Against the
    # engagement target of 4 it would be a gap; against its own target it is not.
    exempt = answers[0]
    c.patch(f"/zt/answers/{exempt['id']}", headers=h, json={"target_stage": 3})
    c.post(f"/zt/assessments/{a['id']}/approve", headers=h)

    total = len(answers)
    fin = c.post(f"/zt/services/{svc_id}/deliverables/finalize", headers=h)
    assert fin.status_code == 201, fin.text

    from app.models.deliverable import Deliverable

    eng = create_engine(os.environ["DATABASE_URL"], future=True)
    with sessionmaker(bind=eng, future=True)() as s:
        summary = s.get(Deliverable, uuid.UUID(fin.json()["id"])).summary or ""

    # Every capability is a gap at S4 EXCEPT the exempt one, so the count is
    # total - 1. Drop `targets=` and it becomes `total`.
    assert (
        f"{total - 1} gap(s) at target S4" in summary
    ), f"the per-capability target was ignored; expected {total - 1} gaps: {summary!r}"


# --- #79 — the CSF twin ----------------------------------------------------


@pytest.mark.unit
def test_csf_export_and_home_card_agree_with_the_dashboard_on_the_target(app_client) -> None:
    """All three CSF surfaces must report the same gap count.

    The dashboard reads the client's intake tier; the exporter and the /home
    value-loop card both used the engine default. For a client whose tier is not
    3 that is zero versus everything, on adjacent screens.
    """
    c = app_client
    admin = _register(c, "admin@example.com")
    client = _register(c, "client@example.com")
    h = {"Authorization": f"Bearer {admin['tokens']['access_token']}"}
    client_id = client["user"]["client_id"]

    svc_id = c.post("/csf/services", headers=h, json={"kind": "nist_csf", "title": "CSF"}).json()[
        "id"
    ]
    _attach_intake_target(svc_id, csf_tier=4)

    a = c.post(f"/csf/services/{svc_id}/assessments", headers=h).json()
    # Tier 3: zero gaps against the default 3, all of them against the client's 4.
    for ans in a["answers"]:
        c.patch(f"/csf/answers/{ans['id']}", headers=h, json={"maturity_tier": 3})
    c.post(f"/csf/assessments/{a['id']}/approve", headers=h)
    fin = c.post(f"/csf/services/{svc_id}/deliverables/finalize", headers=h)
    assert fin.status_code == 201, fin.text
    c.post(f"/csf/deliverables/{fin.json()['id']}/release", headers=h)

    c.headers["X-Client-Id"] = client_id
    ch = {"Authorization": f"Bearer {client['tokens']['access_token']}"}
    dash = c.get(f"/clients/{client_id}/csf/{svc_id}/dashboard", headers=ch).json()
    assert dash["target_tier"] == 4 and dash["target_tier_source"] == "client"
    assert dash["total_gap_count"] > 0

    summary = c.get(f"/clients/{client_id}/value-summary", headers=ch).json()
    assert summary["csf_gap_count"] == dash["total_gap_count"], (
        "the /home card and the dashboard disagree about the same assessment: "
        f"card={summary['csf_gap_count']} dashboard={dash['total_gap_count']}"
    )

    from app.models.deliverable import Deliverable

    eng = create_engine(os.environ["DATABASE_URL"], future=True)
    with sessionmaker(bind=eng, future=True)() as s:
        deliv = s.get(Deliverable, uuid.UUID(fin.json()["id"]))
        doc_summary = deliv.summary or ""
    # Name the TARGET. `str(dash["total_gap_count"]) in doc_summary` looked like
    # a real check and was not: this fixture scores all 106 subcategories, so the
    # summary reads "106/106 subcategories scored; 0 gap(s) at target T3" with
    # the fix reverted — and "106" is in it, via the coverage fraction. It stayed
    # green either way, which is the #72 shape in the test written to close #79.
    assert (
        f"{dash['total_gap_count']} gap(s) at target T4" in doc_summary
    ), f"the released document disagrees with the dashboard: {doc_summary!r}"


@pytest.mark.unit
def test_csf_export_states_the_true_gap_total_when_it_truncates(app_client) -> None:
    """#75's CSF twin — filed against ZT, present identically here.

    CSF truncates through the same `analyze_gaps` to the same DEFAULT_TOP_N of
    20. Fixing the ZT renderers and leaving these three would have reproduced
    the half-fix that #79 exists to document — and change #79 makes it worse,
    not better: raising the target from the engine default to the client's tier
    INCREASES the gap count, so more is hidden precisely where the disclosure
    started mattering more.

    `test_csf_exporters.py` asserted the truncation as intended behaviour
    (`assert ws.max_row == 21` over a 106-gap fixture) without ever asking
    whether the document said so.
    """
    import io

    from openpyxl import load_workbook

    c = app_client
    admin = _register(c, "admin@example.com")
    h = {"Authorization": f"Bearer {admin['tokens']['access_token']}"}

    svc_id = c.post("/csf/services", headers=h, json={"kind": "nist_csf", "title": "CSF"}).json()[
        "id"
    ]
    a = c.post(f"/csf/services/{svc_id}/assessments", headers=h).json()
    for ans in a["answers"]:
        c.patch(f"/csf/answers/{ans['id']}", headers=h, json={"maturity_tier": 1})
    c.post(f"/csf/assessments/{a['id']}/approve", headers=h)

    fin = c.post(f"/csf/services/{svc_id}/deliverables/finalize", headers=h)
    assert fin.status_code == 201, fin.text

    total = c.get(f"/csf/services/{svc_id}/gap-analysis", headers=h).json()["total_gap_count"]
    assert total > 20, "fixture must actually exercise truncation, or this proves nothing"

    art_id = fin.json()["xlsx_artifact_id"]
    raw = c.get(f"/artifacts/{art_id}/download", headers=h).content
    ws = load_workbook(io.BytesIO(raw))["Gap Plan"]
    blob = " ".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
    )

    shown = ws.max_row - 2  # minus the caption row AND the header
    assert shown < total, f"fixture did not truncate: shown={shown} total={total}"
    assert str(total) in blob, (
        f"the CSF Gap Plan shows {shown} of {total} gaps and never states the total — "
        "a client cannot tell anything was omitted"
    )


@pytest.mark.unit
def test_zt_home_card_agrees_with_the_released_report_on_the_target(app_client) -> None:
    """The ZT twin of the /home card fix — `_zt_gap_total` was left behind.

    `_csf_gap_total` was given the client's tier; `_zt_gap_total` twelve lines
    below it kept passing `targets` alone, so every capability with no per-row
    override fell back to DEFAULT_TARGET_STAGE. A client on stage 4 with
    everything scored at 3 reads "0 gaps" on the card and 37 in the report they
    were sent — #79's exact symptom, in the service #73 was filed against.

    There is no ZT client dashboard gap count to compare against (it publishes
    none), so the released document is the other surface.
    """
    c = app_client
    admin = _register(c, "admin@example.com")
    client = _register(c, "client@example.com")
    h = {"Authorization": f"Bearer {admin['tokens']['access_token']}"}
    client_id = client["user"]["client_id"]

    svc_id = c.post(
        "/zt/services", headers=h, json={"kind": "zero_trust_cisa", "title": "ZT"}
    ).json()["id"]
    _attach_intake_target(svc_id, zt_stage=4)

    a = c.post(f"/zt/services/{svc_id}/assessments", headers=h).json()
    for ans in a["answers"]:
        c.patch(f"/zt/answers/{ans['id']}", headers=h, json={"maturity_stage": 3})
    c.post(f"/zt/assessments/{a['id']}/approve", headers=h)
    fin = c.post(f"/zt/services/{svc_id}/deliverables/finalize", headers=h)
    assert fin.status_code == 201, fin.text
    c.post(f"/zt/deliverables/{fin.json()['id']}/release", headers=h)

    expected = len(a["answers"])  # every capability is a gap at S4

    c.headers["X-Client-Id"] = client_id
    ch = {"Authorization": f"Bearer {client['tokens']['access_token']}"}
    summary = c.get(f"/clients/{client_id}/value-summary", headers=ch).json()
    assert summary["zt_gap_count"] == expected, (
        "the /home card and the released report disagree about the same "
        f"assessment: card={summary['zt_gap_count']} report={expected}"
    )
